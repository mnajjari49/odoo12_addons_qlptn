# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from odoo.tools import pycompat
import logging

import datetime
import base64
from io import BytesIO
from operator import itemgetter
from ast import literal_eval

from .mailmerge import MailMerge
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

_logger = logging.getLogger(__name__)


def alter_mapped(records, func):
    """Specially made to extract data because odoo mapped function removes duplicates"""
    if not isinstance(records, models.BaseModel):
        return []
    if isinstance(func, pycompat.string_types):
        recs = records
        for name in func.split('.'):
            sub_func = itemgetter(name)
            try:
                new_res = []
                for rec in recs:
                    new_res += [sub_func(sub_rec) for sub_rec in rec]
                recs = new_res
            except KeyError:
                raise ValidationError(_("Something's wrong with the code: %s" % name))
        return recs
    else:
        return records


def num2words_vnm(num):
    under_20 = ['không', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín', 'mười', 'mười một',
                'mười hai', 'mười ba', 'mười bốn', 'mười lăm', 'mười sáu', 'mười bảy', 'mười tám', 'mười chín']
    tens = ['hai mươi', 'ba mươi', 'bốn mươi', 'năm mươi', 'sáu mươi', 'bảy mươi', 'tám mươi', 'chín mươi']
    above_100 = {100: 'trăm', 1000: 'nghìn', 1000000: 'triệu', 1000000000: 'tỉ'}

    if num < 20:
        return under_20[num]

    elif num < 100:
        under_20[1], under_20[5] = 'mốt', 'lăm'  # thay cho một, năm
        result = tens[num // 10 - 2]
        if num % 10 > 0:  # nếu num chia 10 có số dư > 0 mới thêm ' ' và số đơn vị
            result += ' ' + under_20[num % 10]
        return result

    else:
        unit = max([key for key in above_100.keys() if key <= num])
        result = num2words_vnm(num // unit) + ' ' + above_100[unit]
        if num % unit != 0:
            if num > 1000 and num % unit < unit/10:
                result += ' không trăm'
            if 1 < num % unit < 10:
                result += ' linh'
            result += ' ' + num2words_vnm(num % unit)
    return result.capitalize()


class TemplateCreation(models.Model):
    _name = 'temp.creation'
    _description = 'MS Template'

    name = fields.Char("Template's name", copy=False, required=True, default='/')
    main_model = fields.Many2one('ir.model', string="Model")
    max_samples = fields.Integer('Maximum samples', default=3, help="Maximum samples for testing.")
    template = fields.Binary('Template', required=True)
    template_name = fields.Char('Template')
    template_extension = fields.Char('Extension', compute='_get_file_extension')
    access_groups = fields.Many2many('res.groups', string='Access groups', help='Groups that can access the wizard action.')
    reference = fields.Char('Reference', help='Used for python referring')

    export_horizontally = fields.Boolean('Export horizontally', help="Check if you want to export records horizontally. Does not support words.")
    wizard_action = fields.Many2one(comodel_name='ir.actions.act_window', string='Wizard action', compute='_check_wizard_action')
    chosen_fields = fields.One2many('fields.line', 'line', 'Chosen fields', copy=True)
    debug = fields.Boolean('God mode')
    all_in_one = fields.Boolean('All in one',
                                help='Excel: Check if you want to export all records in one sheet, leave uncheck if you want to export record in separate sheet, \n'
                                     'Words: Check if you want to export all records in one template, leave uncheck if you want to export each record as a template, separate by page break.')
    export_current_datetime = fields.Boolean('Export current time', help='Export current time to template')
    current_datetime_format = fields.Char('Time format', help='Use python strftime format')
    current_datetime_positions = fields.Char('Time positions', help='Cell position in excel, \n'
                                                                    'Mergefield in words, \n'
                                                                    'Can be more than one, separate by comma.')
    current_datetime_format_2 = fields.Char('Time format 2', help='Use python strftime format')
    current_datetime_positions_2 = fields.Char('Time positions 2', help='Cell position in excel, \n'
                                                                        'Mergefield in words, \n'
                                                                        'Can be more than one, separate by comma.')
    export_user_name = fields.Boolean('Export user name', help="Export current user's name")
    export_user_position = fields.Char('User name position', help='Cell position in excel, \n'
                                                                  'Mergefield in words, \n'
                                                                  'Can be more than one, separate by comma.')
    # fields for excel options
    max_col = fields.Char('Hide to col',
                          help='When exporting, empty columns from last column with value to this column will be hidden, only work with one x2many field.')
    max_row = fields.Integer('Hide to row',
                             help='When exporting, empty rows from last row with value to this row will be hidden, only work with one x2many field.')
    sheet_name = fields.Boolean('Differ sheet name', help="Use record's display name as sheet name.")
    grp_by_field = fields.Many2one('ir.model.fields', 'Group by field')
    grp_by_cell = fields.Char('Group by - start cell', help='Choose the cell to write group by - field on, normally one row above the record exporting row')
    # fields for words options
    table_marks = fields.Char('Table marks', help="List all marks of first table here, separate by comma.")
    table2_marks = fields.Char('2nd Table marks', help="List all marks of second table here, separate by comma.")

    _sql_constraints = [
        ('name_model_unique', 'unique (name, main_model)', "Template's name must be unique per model."),
    ]

    @api.onchange('export_current_datetime')
    def _onchange_export_current_datetime(self):
        if self.export_current_datetime:
            self.current_datetime_format = 'Ngày %d tháng %m năm %Y'
        else:
            self.current_datetime_format = False
            self.current_datetime_format_2 = False
            self.current_datetime_positions = False
            self.current_datetime_positions_2 = False

    @api.one
    @api.depends('template_name')
    def _get_file_extension(self):
        if self.template_name:
            self.template_extension = str(self.template_name).rsplit('.', 1)[1]

    @api.one
    @api.constrains('template_name')
    def _constrain_extension(self):
        if not self.template or self.template_extension not in ['doc', 'docx', 'xls', 'xlsx']:
            raise ValidationError(_('Please upload docx, xls or xlsx file for template.'))

    @api.one
    @api.constrains('max_col', 'grp_by_cell', 'current_datetime_positions')
    def _constrain_excel_cell(self):
        if self.template_extension in ['xls', 'xlsx']:
            if self.max_col and not self.max_col.isalpha():
                raise ValidationError(_("Hide to col's value must be string."))
            if self.grp_by_cell and not self.grp_by_cell.strip().rstrip('0123456789').isalpha():
                raise ValidationError(_("Invalid 'Group by - start cell' value format."))
            if self.current_datetime_positions:
                for cell_position in self.current_datetime_positions.split(','):
                    if not cell_position.strip().rstrip('0123456789').isalpha():
                        raise ValidationError(_("Invalid 'Current datetime position' value format."))
            if self.current_datetime_positions_2:
                for cell_position in self.current_datetime_positions_2.split(','):
                    if not cell_position.strip().rstrip('0123456789').isalpha():
                        raise ValidationError(_("Invalid 'Current datetime position' value format."))

    @api.one
    def write(self, vals):
        res = super(TemplateCreation, self).write(vals)
        if self.wizard_action:
            if vals.get('access_groups'):
                self.wizard_action.groups_id = [(6, 0, self.access_groups.ids)]
            if vals.get('name'):
                self.wizard_action.name = self.name
        return res

    @api.multi
    def unlink(self):
        for template in self:
            if template.wizard_action:
                raise UserError(_('In order to delete a template, you must remove its action.'))
        return super(TemplateCreation, self).unlink()

    @api.multi
    @api.depends('main_model')
    def _check_wizard_action(self):
        for record in self:
            exist = self.env['ir.actions.act_window'].search([('name', '=', record.name), ('binding_model_id', '=', record.main_model.id)])
            if exist:
                record.wizard_action = exist.id

    @api.multi
    def create_wizard_action(self):
        if not self.wizard_action:
            self.env['ir.actions.act_window'].create({
                'name': self.name,
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'tree,form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'binding_model_id': self.main_model.id,
                'groups_id': [(6, 0, self.access_groups.ids)],
                'context': {'default_template_id': self.id}})
            return {'type': 'ir.actions.client',
                    'tag': 'reload'}
        else:
            self.env['ir.actions.act_window'].search([('name', '=', self.name),
                                                      ('binding_model_id', '=', self.main_model.id)]).unlink()
            return {'type': 'ir.actions.client',
                    'tag': 'reload'}

    @api.onchange('template_name')
    def onchange_temp_name(self):
        if self.template_name and (not self.name or self.name == '/'):
            self.name = str(self.template_name).rsplit('.', 1)[0]

    @api.onchange('all_in_one')
    def onchange_all_in_one(self):
        if not self.all_in_one:
            self.grp_by_field = False
            self.grp_by_cell = False
        else:
            self.sheet_name = False

    @api.multi
    def get_attachment_view(self):
        attachment_action = self.env.ref('base.action_attachment')
        action = attachment_action.read()[0]
        action['context'] = {'default_res_model': self._name, 'default_res_id': self.ids[0]}
        action['domain'] = str(['&', ('res_model', '=', self._name), ('res_id', 'in', self.ids)])
        action['search_view_id'] = (self.env.ref('base.view_attachment_search').id,)
        return action

    def get_field_value(self, records, code, replacement_format):
        self.ensure_one()
        try:
            r_f_dict = literal_eval(replacement_format) if replacement_format else {}
        except:
            raise ValidationError(_('Wrong value for Replace & format field'))
        res = []
        values = []
        for record in records:
            values += alter_mapped(record, code)
        for v in values:
            if isinstance(type(v), models.MetaModel):
                v = ', '.join(sv.display_name for sv in v)
            elif isinstance(v, datetime.date) and r_f_dict.get('time_format'):
                v = v.strftime(r_f_dict['time_format'])
            elif isinstance(v, int) or isinstance(v, float):
                if r_f_dict.get('n2w'):
                    v = num2words_vnm(int(v))
                elif r_f_dict.get('to_int'):
                    v = '{0:,}'.format(int(v))
            res.append(r_f_dict.get(str(v), v))
        if r_f_dict.get('to_upper'):
            res = [str(v).upper() for v in res]
        if r_f_dict.get('str_format'):
            res = [r_f_dict['str_format'] % v for v in res]
        if r_f_dict.get('all_to_str'):
            res = [', '.join(str(v) for v in res)]
        return res

    def export_excel(self, records):
        self.ensure_one()
        decode = base64.b64decode(self.template)
        wb = load_workbook(BytesIO(decode))
        working_ws = [wb.active]
        mixin_list = records
        if not self.all_in_one:
            if len(records) > 1:
                wb.active.title = records[0].display_name.translate({ord(i): None for i in '\/*[]:?'}) if self.sheet_name else wb.active.title
                for i in range(1, len(records)):
                    new_ws = wb.copy_worksheet(wb.active)
                    if self.sheet_name:
                        new_ws.title = records[i].display_name.translate({ord(i): None for i in '\/*[]:?'})
                    else:
                        new_ws.title = wb.active.title + ' ' + str(i + 1)
                    working_ws.append(new_ws)
        # format mixin_list to export record with grp_by_fields: [grp_by_value_1, record 1, record 2, value 2, record 3, record 4]
        elif self.grp_by_field:
            grp_cell_col_text = self.grp_by_cell.strip().rstrip('0123456789')
            grp_start_col = column_index_from_string(grp_cell_col_text)
            grp_start_row = int(self.grp_by_cell.strip()[len(grp_cell_col_text):])
            mapped_string = self.grp_by_field.name
            grp_by_list = records.mapped(self.grp_by_field.name)
            if isinstance(type(grp_by_list), models.MetaModel):
                mapped_string += '.' + 'display_name'
                grp_by_list = grp_by_list.mapped('display_name')  # to export display name to excel
            mixin_list = []
            for value in grp_by_list:
                mixin_list.append(value)
                current_records = records.filtered(lambda r: r.mapped(mapped_string) == [value])
                for record in current_records:
                    mixin_list.append(record)  # odoo .sorted only accept string key so this is the other way around to sort records by a relation field

        for num, ws in enumerate(working_ws):
            # export current datetime with defined format to defined cells:
            if self.export_current_datetime:
                position_list = [position.strip() for position in self.current_datetime_positions.split(',')]
                current_datetime_value = fields.datetime.now().strftime(self.current_datetime_format)
                for position in position_list:
                    ws[position] = current_datetime_value
                if self.current_datetime_positions_2:
                    position_list_2 = [position.strip() for position in self.current_datetime_positions_2.split(',')]
                    current_datetime_value_2 = fields.datetime.now().strftime(self.current_datetime_format_2)
                    for position2 in position_list_2:
                        ws[position2] = current_datetime_value_2
            if self.export_user_name:
                user_position_list = [position.strip() for position in self.export_user_position.split(',')]
                for u_position in user_position_list:
                    ws[u_position] = self.env.user.name
            mixin_list = mixin_list if self.all_in_one else records[num]
            distance = 0  # the distance from start row / col to the row / col of current record
            max_cell = 0  # hold the max row / col with value exported, so row / col after it will be hidden
            for num2, item in enumerate(mixin_list):
                if not isinstance(type(item), models.MetaModel):  # if item is a grp_by_field value
                    if num2 > 0:
                        distance += 1  # to write group by value in between records
                    if self.export_horizontally:
                        ws.cell(grp_start_row, grp_start_col + distance).value = item
                    else:
                        ws.cell(grp_start_row + distance, grp_start_col).value = item
                else:
                    temp_distance = 0  # temporary distance in one record: the max len of record's x2many fields
                    for line in self.chosen_fields:
                        export_res = self.get_field_value(item, line.python_code, line.replacement_format)
                        temp_distance = max(len(export_res), temp_distance)
                        for cell_location in line.temp_mark.split(','):
                            cell_col_text = cell_location.strip().rstrip('0123456789')
                            start_col = column_index_from_string(cell_col_text)
                            start_row = int(cell_location.strip()[len(cell_col_text):])
                            for n, v in enumerate(export_res):
                                if self.export_horizontally:
                                    ws.cell(row=start_row, column=start_col + distance + n).value = v
                                    if (len(export_res) > 1 or self.all_in_one) and self.max_col:
                                        max_cell = start_col + distance + temp_distance + 1
                                else:
                                    ws.cell(row=start_row + distance + n, column=start_col).value = v
                                    if (len(export_res) > 1 or self.all_in_one) and self.max_row > 0:
                                        max_cell = start_row + distance + temp_distance + 1
                    distance += temp_distance
            if max_cell > 0:
                if self.export_horizontally and max_cell < column_index_from_string(self.max_col):
                    ws.column_dimensions.group(start=get_column_letter(max_cell), end=self.max_col, hidden=True, outline_level=0)
                elif not self.export_horizontally and max_cell < self.max_row:
                    ws.row_dimensions.group(start=max_cell, end=self.max_row, hidden=True, outline_level=0)

        # wb.security = WorkbookProtection()
        # wb.security.workbookPassword = '1'
        # wb.security.lockStructure = True
        # ws.protection.password = '1'

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)

        report = base64.encodebytes((fp.read()))
        fp.close()
        return report

    def export_words(self, records):
        self.ensure_one()
        decode = base64.b64decode(self.template)
        doc = MailMerge(BytesIO(decode))
        """self.get_field_value() return a list of values of A field, similar to a columns of data
        while mailmerge table format is: {'col1': [{field1: value1, field2: value2}, {field1: value3, field2: value4}]
        with each dict as a row, this func will return table data for mailmerge"""
        def get_table_rows(keys, columns):  # keys is list of string, columns is list of list
            table_rows = []
            columns_max_len = max([len(col) for col in columns])
            for col in columns:  # standardize columns len
                if len(col) < columns_max_len:
                    col += [' '] * (columns_max_len - len(col))
            for i in range(len(columns[0])):
                pair = []
                for k in range(len(keys)):
                    pair.append((keys[k], columns[k][i]))
                table_rows.append(dict(pair))
            return table_rows

        records_data = {}
        if self.export_current_datetime:  # export current datetime with defined format to defined position
            position_list = [position.strip() for position in self.current_datetime_positions.split(',')]
            current_datetime_value = fields.datetime.now().strftime(self.current_datetime_format)
            for position in position_list:
                records_data[position] = current_datetime_value
            if self.current_datetime_positions_2:
                position_list_2 = [position.strip() for position in self.current_datetime_positions_2.split(',')]
                current_datetime_value_2 = fields.datetime.now().strftime(self.current_datetime_format_2)
                for position2 in position_list_2:
                    records_data[position2] = current_datetime_value_2
        if self.export_user_name:
            user_position_list = [position.strip() for position in self.export_user_position.split(',')]
            for u_position in user_position_list:
                records_data[u_position] = self.env.user.name

        if not self.all_in_one:
            data_list = []  # final format: [record_data 1, record_data 2...]
            table1_keys = [key.strip() for key in self.table_marks.split(',')] if self.table_marks else []
            table2_keys = [key.strip() for key in self.table2_marks.split(',')] if self.table2_marks else []
            for record in records:
                record_data = records_data  # final format: {mark1: value1, mark2: value2, col1: [row1, row2]}

                record_table1_columns = []  # hold table column values, will be combined with keys to return table data
                record_table2_columns = []
                record_table1_rows = []
                record_table2_rows = []
                for line in self.chosen_fields:
                    export_res = [str(res) for res in self.get_field_value(record, line.python_code, line.replacement_format)]
                    for temp_mark in line.temp_mark.split(','):
                        if temp_mark.strip() in table1_keys:
                            record_table1_rows.append(temp_mark.strip())  # reordering keys
                            record_table1_columns.append(export_res)
                        elif temp_mark.strip() in table2_keys:
                            record_table2_rows.append(temp_mark.strip())
                            record_table2_columns.append(export_res)
                        else:
                            record_data[temp_mark.strip()] = ', '.join(export_res)
                if record_table1_columns:
                    record_data[table1_keys[0]] = get_table_rows(record_table1_rows, record_table1_columns)
                if record_table2_columns:
                    record_data[table2_keys[0]] = get_table_rows(record_table2_rows, record_table2_columns)
                data_list.append(record_data)
            doc.merge_templates(data_list, separator='page_break')

        else:
            records_keys = []
            for temp_mark in self.chosen_fields.mapped('temp_mark'):  # striping multiple marks in one temp_mark
                records_keys += [mark.strip() for mark in temp_mark.split(',')]
            records_columns = [[] for i in range(len(records_keys))]
            # working record by record
            for record in records:
                for num, line in enumerate(self.chosen_fields):
                    records_columns[num] += [str(res) for res in self.get_field_value(record, line.python_code, line.replacement_format)]
                max_len = max([len(col) for col in records_columns])
                for col in records_columns:   # standardize column length
                    if len(col) < max_len:
                        col += [' '] * (max_len - len(col))
            records_data[records_keys[0]] = get_table_rows(records_keys, records_columns)
            doc.merge(**records_data)
        fp = BytesIO()
        doc.write(fp)
        doc.close()
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        return report

    def export_records_data(self, records):
        self.ensure_one()
        if self.template_extension in ['xls', 'xlsx']:
            return self.export_excel(records)
        elif self.template_extension in ['doc', 'docx']:
            return self.export_words(records)

    @api.multi
    def sample_report(self):
        self.env['ir.attachment'].create({
            'name': 'Another weird report',
            'datas_fname': self.template_name,
            'datas': self.export_records_data(self.env[self.main_model.model].search([], limit=self.max_samples or 3)),
            'res_model': 'temp.creation',
            'res_id': self.id})
        return self.get_attachment_view()

    @api.model
    def clean_attachments(self):
        self.env['ir.attachment'].search([('res_model', '=', 'temp.creation'), ('public', '=', True)]).unlink()


class ChosenFields(models.Model):
    _name = 'fields.line'

    line = fields.Many2one('temp.creation')
    main_field = fields.Many2one('ir.model.fields', string='Main field')
    sub_model = fields.Many2one('ir.model', string='Sub model', compute='_get_sub_model')
    sub_field = fields.Many2one('ir.model.fields', string='Sub field')
    field_value = fields.Char('Field value', compute='_get_value')
    temp_mark = fields.Char('Position', help="Cell position for excel; \n"
                                             "Mergefield for words - press Alt-N-Q-F-M-M to insert one; \n"
                                             "Separate by comma if you want to input multiple marks")
    python_code = fields.Char('Python code')
    replacement_format = fields.Char('Replace & format',
                                     help="Define replacement values and format by a dictionary. \n"
                                          "Available command: \n"
                                          "as_str: export all result of the field to one string. \n"
                                          "is_upper: convert the text to uppercase. \n"
                                          "n2w: convert the number to words in Vietnamese. \n"
                                          "time_format: set format for datetime object. \n"
                                          "And any value you need to replace. \n"
                                          "Example: {'False': ' ', 'male': 'Mr.', 'female': 'Ms', 'is_upper': True, 'n2w': True, 'time_format': '%d - %m - %Y'}")

    _sql_constraints = [
        ('mark_uniq', 'unique(temp_mark, line)', 'Mark in template must be unique!')]

    @api.one
    @api.constrains('replacement_format')
    def _constrain_replacement_format(self):
        if self.replacement_format:
            replacement_dict = {}
            try:
                replacement_dict = literal_eval(self.replacement_format)
            except:
                raise ValidationError(_('Wrong format for Replace & format dictionary.'))
            if replacement_dict:
                for key, value in replacement_dict.items():
                    if not isinstance(key, pycompat.string_types) or (not isinstance(value, pycompat.string_types) and not isinstance(value, bool)):
                        raise ValidationError(_('Replacement value must be string.'))

    @api.onchange('main_field')
    def onchange_main_field(self):
        self.sub_field = False
        self.python_code = self.main_field.name

    @api.onchange('sub_field')
    def onchange_sub_field(self):
        if self.main_field and not self.sub_field:
            self.python_code = self.python_code.split('.')[0]
        elif self.sub_field:
            self.python_code = self.python_code.split('.')[0] + '.' + self.sub_field.name

    # This is just to show values
    @api.one
    @api.depends('python_code', 'main_field', 'sub_field', 'line.main_model', 'line.max_samples', 'replacement_format')
    def _get_value(self):
        if self.line.main_model:
            records = self.env[self.line.main_model.model].search([], limit=self.line.max_samples or 3)
            list_value = self.line.get_field_value(records, self.python_code, self.replacement_format)
            self.field_value = ', '.join(str(i) for i in list_value)

    @api.one
    @api.depends('main_field')
    def _get_sub_model(self):
        if self.main_field.ttype in ['many2one', 'one2many', 'many2many']:
            self.sub_model = self.env['ir.model'].search([('model', '=', self.main_field.relation)])

    @api.constrains('temp_mark')
    def _constrain_excel_temp_mark(self):
        for record in self:
            if record.line.template_extension in ['xls', 'xlsx']:
                for mark in record.temp_mark.split(','):
                    if not mark.strip().rstrip('0123456789').isalpha():
                        raise ValidationError(_("Invalid Mark value format: %s" % mark))

    # depreciated python-docx export function
    # def export_words(self, records):
    #     self.ensure_one()
    #     decode = base64.b64decode(self.template)
    #     doc = Document(BytesIO(decode))
    #     if len(records) == 1:
    #         for p in doc.paragraphs:
    #             for line in self.chosen_fields:
    #                 if line.temp_mark in p.text:
    #                     for run in p.runs:
    #                         if run.text == line.temp_mark:
    #                             export_res = self.get_field_value(records, line.python_code, line.as_str)
    #                             run.text = ', '.join(str(i) for i in export_res)
    #         for t in doc.tables:
    #             for r in range(len(t.rows)):
    #                 for c in range(len(t.columns)):
    #                     for line in self.chosen_fields:
    #                         if line.temp_mark in t.cell(r, c).text:
    #                             export_res = self.get_field_value(records, line.python_code, line.as_str)
    #                             for i, v in enumerate(export_res):
    #                                 if self.export_horizontally:
    #                                     t.cell(r, c + i).paragraphs[0].runs[-1].text = str(v)
    #                                 else:
    #                                     t.cell(r + i, c).paragraphs[0].runs[-1].text = str(v)
    #     else:
    #         for t in doc.tables:
    #             for r in range(len(t.rows)):
    #                 for c in range(len(t.columns)):
    #                     for line in self.chosen_fields:
    #                         export_res = self.get_field_value(records, line.python_code, line.as_str)
    #                         for i, v in enumerate(export_res):
    #                             if line.temp_mark in t.cell(r, c).text:
    #                                 if self.export_horizontally:
    #                                     t.cell(r, c + i).paragraphs[0].runs[-1].text = str(v)
    #                                 else:
    #                                     t.cell(r + i, c).paragraphs[0].runs[-1].text = str(v)
    #
    #     fp = BytesIO()
    #     doc.save(fp)
    #     fp.seek(0)
    #     report = base64.encodebytes((fp.read()))
    #     fp.close()
    #     return report
